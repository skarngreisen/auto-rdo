# -*- coding: utf-8 -*-
import openpyxl, re
from datetime import datetime

def fmt_horas(v):
    if v is None or str(v).startswith('='): return ''
    s = str(v)
    days = re.search(r'(\d+)\s*days?', s)
    t = re.search(r'(\d+):(\d+):(\d+)', s)
    d = int(days.group(1)) if days else 0
    if t: return f"{d}d {int(t.group(1)):02d}:{int(t.group(2)):02d}h"
    return s[:20]

def safe(v):
    if v is None: return ''
    if isinstance(v, datetime): return v.strftime('%d/%m/%Y')
    return str(v)[:45]

wb = openpyxl.load_workbook(
    "DH/Sistema Controle Horas Paradas PARC DAS ARTES concluido.xlsx",
    data_only=True
)
# Sheets by index: 0=Fevereiro, 1=Marco, 2=Abril, 3=Maio, 9=DASHBOARD, 
# 10=DISTRIBUICAO, 11=DASHBOARD MENSAL, 12=DASHBOARD GERAL EXECUCAO

print("# Dados Consolidados - Parc das Artes\n")

# Dashboard Geral de Execucao
ws = wb.worksheets[12]
print("## Projeto\n")
for i in range(3, 16):
    vals = [safe(ws.cell(row=i, column=c).value) for c in range(1, 9)]
    vals = [v for v in vals if v]
    if vals: print('  ' + ' | '.join(vals))

# Distribuicao Horas Paradas
ws = wb.worksheets[10]
print("\n## Horas Paradas por Categoria\n")
for i in range(3, 10):
    cat = safe(ws.cell(row=i, column=1).value)
    hp = fmt_horas(ws.cell(row=i, column=2).value)
    hdh = fmt_horas(ws.cell(row=i, column=5).value)
    if cat: print(f"  {cat:<30} Parc: {hp:<12} DH: {hdh}")

# Dashboard
ws = wb.worksheets[9]
print("\n## Resumo do Dashboard\n")
for i in range(4, 19):
    vals = [safe(ws.cell(row=i, column=c).value) for c in range(1, 7)]
    vals = [v for v in vals if v]
    if vals: print('  ' + ' | '.join(vals))

# Produtividade Mensal
ws = wb.worksheets[11]
print("\n## Produtividade Mensal\n")
for i in range(3, 8):
    vals = [safe(ws.cell(row=i, column=c).value) for c in range(1, 4)]
    vals = [v for v in vals if v]
    if vals: print('  ' + ' | '.join(vals))

# Equipamentos - resumo
ws = wb.worksheets[4]
print("\n## Equipamentos\n")
for i in range(3, 8):
    vals = [safe(ws.cell(row=i, column=c).value) for c in range(1, 5)]
    vals = [v for v in vals if v]
    if vals: print('  ' + ' | '.join(vals))

wb.close()
print("\n---")
print("Validar contra a planilha original antes de usar em reuniao.")
